"""Generate program description + complete code Excel files"""
import os, sys, json
from pathlib import Path
from datetime import datetime
sys.path.insert(0, ".")
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DESKTOP = Path.home() / "Desktop"
PROJ = Path(__file__).parent

# Styles
hdr_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
acc_fill = PatternFill(start_color="4f8ef7", end_color="4f8ef7", fill_type="solid")
light_fill = PatternFill(start_color="e8eaf6", end_color="e8eaf6", fill_type="solid")
title_font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
h1_font = Font(name="Microsoft YaHei", size=14, bold=True, color="1a1a2e")
h2_font = Font(name="Microsoft YaHei", size=12, bold=True, color="4f8ef7")
body_font = Font(name="Microsoft YaHei", size=11)
code_font = Font(name="Consolas", size=10)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

# ==================== File 1: Program Description ====================
wb1 = openpyxl.Workbook()
ws = wb1.active; ws.title = "TreeCut说明"
ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 40; ws.column_dimensions['C'].width = 90
row = 1

def add_title(text, r):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=text); c.font = title_font; c.fill = hdr_fill; c.alignment = center
    ws.row_dimensions[r].height = 40; return r + 1

def add_h1(text, r):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=1, value=text).font = h1_font; ws.cell(row=r, column=1).fill = light_fill
    ws.row_dimensions[r].height = 30; return r + 1

def add_row(label, value, r):
    ws.cell(row=r, column=1, value=label).font = h2_font; ws.cell(row=r, column=1).fill = light_fill; ws.cell(row=r, column=1).border = thin_border
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value=value).font = body_font; ws.cell(row=r, column=2).fill = PatternFill(); ws.cell(row=r, column=2).border = thin_border; ws.cell(row=r, column=2).alignment = wrap
    ws.row_dimensions[r].height = max(28, 15 * (len(str(value)) // 70 + 1)); return r + 1

row = add_title("TreeCut v10.3 - AI Video Editor", row); row += 1
row = add_h1("What It Is", row)
row = add_row("Product", "AI-powered semi-automatic video editing tool for Xiaohongshu kitchen island/home industry. Input keywords -> AI generates scripts + voiceover + matches footage -> JianYing Pro draft files.", row)
row = add_row("Version", "v10.3 (2026-06-11)", row)
row = add_row("Tech Stack", "Python 3.12 + DeepSeek API + Edge TTS + Qwen3-VL-4B + Florence-2 + CLIP ViT-L + YOLOv8n + SenseVoice + Faster-Whisper + FAISS + BGE-M3 + Ollama", row)
row = add_row("Code Scale", "42 Python modules, 55 files total", row)
row += 1
row = add_h1("Core Features", row)
features = [
    ("AI Script Generation", "Input keyword -> DeepSeek generates 3-segment Xiaohongshu copy (hook+sell+CTA), 200+ industry terms knowledge base"),
    ("AI Voiceover", "Edge TTS (Xiaoyi voice), 1.1x speed, 1254 protected words prevent word splitting, 5-layer text cleaning"),
    ("Smart Material Matching", "3-tier: FAISS vector search -> Knowledge base rules (200+ terms) -> Filename keyword matching"),
    ("4 Vision Models", "Qwen3-VL-4B (local) + Florence-2 (local) + CLIP ViT-L + YOLOv8n + KnowledgeBridge"),
    ("Audio Analysis", "SenseVoice (emotion+events, priority) + Faster-Whisper (backup)"),
    ("Full-Disk Video Search", "Scan all drives, tree display, one-click batch generation from selected folders"),
    ("Frame Annotation", "Extract keyframes -> 4-model recognition -> manual edit -> feedback learning closed loop"),
    ("Batch Production", "Excel import / manual input / folder selection 3 modes, generate N videos at once"),
    ("Material Library", "35 selling point folders, 15387 indexed segments, FAISS vector index, quality scoring + dedup"),
    ("Self-Learning", "DeepSeek daily analyzes usage records -> generates optimization rules -> auto-applies to models"),
]
for label, val in features: row = add_row(label, val, row)
row += 1
row = add_h1("System Requirements", row)
for l, v in [("OS","Windows 10/11 64-bit"),("Python","3.12"),("RAM","8GB min, 16GB recommended"),("Disk","~20GB (models ~17GB)"),("Network","Required (DeepSeek API + Edge TTS)")]: row = add_row(l, v, row)
row += 1
row = add_h1("Launch Methods", row)
for l, v in [("Desktop","python tree.py / double-click shortcut"),("Web","python tree.py --web -> http://localhost:7860"),("CLI","python tree.py keyword --tts --auto-bgm"),("Setup","python tree.py --setup"),("Status","python tree.py --status")]: row = add_row(l, v, row)
wb1.save(str(DESKTOP / "TreeCut_Program_Description.xlsx"))
print("File 1: TreeCut_Program_Description.xlsx generated")

# ==================== File 2: Complete Code ====================
wb2 = openpyxl.Workbook()
ws2 = wb2.active; ws2.title = "File Index"
ws2.column_dimensions['A'].width = 5; ws2.column_dimensions['B'].width = 45; ws2.column_dimensions['C'].width = 18; ws2.column_dimensions['D'].width = 12; ws2.column_dimensions['E'].width = 60
row2 = 1
for ci, (col, w) in enumerate([("No.",5),("File Path",45),("Module",18),("Lines",12),("Description",60)]):
    c = ws2.cell(row=row2, column=ci+1, value=col); c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF"); c.fill = hdr_fill; c.alignment = center; c.border = thin_border
row2 += 1

# Collect all Python files
py_files = []
for root, dirs, files in os.walk(PROJ):
    dirs[:] = [d for d in dirs if d not in ('__pycache__','.git','02_BGM','03_粗剪输出','04_文案','05_配音','custom_voice','shipin','models','.offload','material_engine_v3')]
    for f in files:
        if f.endswith('.py') or f.endswith('.bat') or f.endswith('.txt') or f.endswith('.json') or f.endswith('.md') or f.endswith('.iss'):
            fp = Path(root)/f; rel = str(fp.relative_to(PROJ))
            mod = str(Path(root).relative_to(PROJ)) if str(Path(root).relative_to(PROJ)) != '.' else 'root'
            try: lines = len(fp.read_text(encoding='utf-8').split('\n'))
            except Exception: lines = 0
            desc = {"核心引擎":"engine","用户界面":"ui","工具模块":"utils","测试":"test","":"config"}.get(os.path.basename(root), os.path.basename(root))
            py_files.append((rel, mod, lines, desc))

py_files.sort(key=lambda x: x[0])
total_lines = 0
for i, (fp, mod, lines, desc) in enumerate(py_files):
    total_lines += lines
    vals = [i+1, fp, mod, lines, desc]
    for ci, val in enumerate(vals):
        c = ws2.cell(row=row2, column=ci+1, value=val); c.font = body_font; c.border = thin_border
        if ci in (0, 3): c.alignment = center
    row2 += 1

row2 += 1; ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
ws2.cell(row=row2, column=1, value=f"Total: {len(py_files)} files, {total_lines:,} lines").font = Font(name="Microsoft YaHei", size=12, bold=True)

# Add code sheets for key files
for fp, mod, lines, desc in py_files:
    if lines > 0 and lines < 2000 and fp.endswith('.py'):
        fpath = PROJ / fp
        try:
            safe = fp.replace('/','_').replace('\\','_').replace('.','_')[:31]
            ws = wb2.create_sheet(title=safe)
            ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 160
            ws.merge_cells('A1:B1'); ws.cell(row=1, column=1, value=f"{fp} | {mod} | {lines} lines | {desc}").font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF"); ws.cell(row=1, column=1).fill = hdr_fill
            ws.cell(row=2, column=1, value="Line").font = h2_font; ws.cell(row=2, column=1).fill = light_fill; ws.cell(row=2, column=1).border = thin_border
            ws.cell(row=2, column=2, value="Code").font = h2_font; ws.cell(row=2, column=2).fill = light_fill; ws.cell(row=2, column=2).border = thin_border
            for ln, text in enumerate(fpath.read_text(encoding='utf-8').split('\n'), 1):
                c1 = ws.cell(row=ln+2, column=1, value=ln); c1.font = Font(name="Consolas", size=9, color="888888"); c1.alignment = Alignment(horizontal="right", vertical="top"); c1.border = thin_border
                c2 = ws.cell(row=ln+2, column=2, value=text); c2.font = code_font; c2.border = thin_border
                if ln % 2 == 0: c1.fill = c2.fill = PatternFill(start_color="fafafa", end_color="fafafa", fill_type="solid")
        except Exception: pass
wb2.save(str(DESKTOP / "TreeCut_Complete_Code.xlsx"))
print(f"File 2: TreeCut_Complete_Code.xlsx ({len(py_files)} files, {total_lines:,} lines)")
print("Done!")
