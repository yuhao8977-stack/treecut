"""
╔══════════════════════════════════════════════════════════════╗
║  [DEDUP] video_editor.pipeline — 视频生成主流程管道             ║
║                                                            ║
║  核心函数: run() / run_multi() / run_batch()               ║
║  素材管理: collect_* / select_clips / match_clips_*        ║
║  验证复盘: validate_draft / build_tts_synced_timeline      ║
╚══════════════════════════════════════════════════════════════╝
"""
import os
import re
import sys
import json
import time
import math
import random
import shutil
import sqlite3
import gc
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict

from core.config import *
from core.copywriter import (
    generate_copy, generate_fallback_copy,
    check_cta_present, append_cta_if_missing,
)
from core.tts import (
    generate_tts_voiceover, get_audio_duration_seconds,
    split_copy_to_subtitles, clean_text_for_tts,
    strip_leading_junk, estimate_tts_duration,
)
# v11.6: 自适应匹配引擎
from core.smart_match_engine import (
    get_adaptive_clips, convert_to_clips, script_hash as make_script_hash,
    record_script_match,
)
from core.draft import (
    JianyingDraftBuilder, save_draft, sec_to_us,
)

# ═══════════════════════════════════════════════════════════════
# 行业知识库 (石材/工艺/五金/风格) — 始终可用
# ═══════════════════════════════════════════════════════════════
from utils.knowledge import KnowledgeBridge
_kb_bridge = KnowledgeBridge()

# AI素材库语义检索 — V3.5 使用SmartMatcher (向量+知识库)
# v11.0: 强制使用FAISS，无降级到纯SQL关键词
try:
    from core.smart_matcher_v3 import get_smart_matcher
    _smart_matcher = get_smart_matcher()
    _HAS_AI_LIBRARY = True
except Exception as e:
    print(f"   [FATAL] SmartMatcher 加载失败: {e}")
    print(f"   向量检索是核心功能，请检查 material_engine_v3/core/smart_matcher.py 是否存在。")
    print(f"   如果 FAISS 索引缺失，请运行: python force_rebuild_faiss.py")
    raise

# ═══════════════════════════════════════════════════════════════
# 素材扫描 — 基础工具函数
# ═══════════════════════════════════════════════════════════════


def list_all_mp4(folder: Path) -> List[Path]:
    """递归列出文件夹下所有 .mp4 文件"""
    mp4s = []
    try:
        for f in folder.rglob("*.mp4"):
            if f.is_file():
                mp4s.append(f)
    except (OSError, PermissionError):
        pass
    mp4s.sort(key=lambda x: x.name)
    return mp4s


def find_closest_folder(keyword: str, base_dir: str) -> Optional[Path]:
    """在 base_dir 下查找最匹配关键词的文件夹"""
    base = Path(base_dir)
    if not base.exists():
        return None
    folders = [d for d in base.iterdir() if d.is_dir()]
    if not folders:
        return None
    kw_clean = re.sub(r'[\[\]\\s]', '', keyword).lower()
    for f in folders:
        if keyword in f.name:
            return f
    for f in folders:
        f_clean = re.sub(r'[\[\]\\s]', '', f.name).lower()
        if kw_clean in f_clean:
            return f
    best_match, best_score = None, 0
    for f in folders:
        f_clean = re.sub(r'[\[\]\\s]', '', f.name).lower()
        score = sum(1 for ch in kw_clean if ch in f_clean)
        if score > best_score:
            best_score, best_match = score, f
    if best_score >= max(1, len(kw_clean) * 0.5):
        return best_match
    return None




# ═══════════════════════════════════════════════════════════════
# 素材收集
# ═══════════════════════════════════════════════════════════════


def collect_effects_mp4s(keyword: str = "", min_count: int = 0,
                         generic_only: bool = False) -> List[Path]:
    """收集效果展示素材"""
    base = Path(EFFECTS_DIR)
    if not base.exists(): return []
    generic, specific = [], []
    for d in sorted(base.iterdir()):
        if not d.is_dir() or "封面图" in d.name: continue
        mp4s = list_all_mp4(d)
        if any(gf in d.name for gf in GENERIC_EFFECTS_FOLDERS):
            generic.extend(mp4s)
        elif keyword:
            specific.extend([m for m in mp4s if keyword in m.name or keyword in d.name])
        else:
            generic.extend(mp4s)
    seen = set()
    result = []
    for f in specific + generic:
        fid = (f.name, f.stat().st_size)
        if fid not in seen:
            seen.add(fid); result.append(f)
    return result


def collect_b_group_mp4s(keyword: str = "") -> List[Path]:
    """从 B组更新视频收集补充素材"""
    if not ENABLE_B_GROUP_MIX: return []
    base = Path(B_GROUP_PATH)
    if not base.exists(): return []
    all_mp4s = []
    for d in sorted(base.iterdir()):
        if not d.is_dir(): continue
        mp4s = list_all_mp4(d)
        if keyword:
            mp4s = [m for m in mp4s if keyword in m.name or keyword in d.name]
        all_mp4s.extend(mp4s)
    all_mp4s.sort(key=lambda x: x.name)
    return all_mp4s


def collect_multi_point_mp4s(keyword: str) -> dict:
    """强制多卖点混合素材收集"""
    result = {"primary": [], "primary_name": "", "other_points": [], "global_show": [], "all_other_mp4s": []}
    folder = find_closest_folder(keyword, SELLING_POINT_DIR)
    if folder:
        result["primary"] = list_all_mp4(folder)
        result["primary_name"] = folder.name
    all_points = list_available_selling_points(min_files=3)
    other_points = [p for p in all_points if p["original_name"] != result["primary_name"]]
    num_other = min(3, len(other_points), random.randint(2, 3))
    if num_other > 0 and other_points:
        for pt in random.sample(other_points, min(num_other, len(other_points))):
            mp4s = list_all_mp4(pt["folder"])
            if mp4s:
                result["other_points"].append({"name": pt["original_name"], "mp4s": mp4s})
                result["all_other_mp4s"].extend(mp4s)
    result["global_show"] = collect_effects_mp4s(keyword)
    return result


def collect_multi_selling_mp4s(keywords: List[str], clips_per_keyword: int = 2) -> List[dict]:
    """从多个卖点文件夹收集素材片段"""
    available = list_available_selling_points()
    if not available: return []
    kw_to_folder = {}
    for kw in keywords:
        folder = find_closest_folder(kw, SELLING_POINT_DIR)
        if folder: kw_to_folder[kw] = folder
    speed_patterns = [1.0, 0.85, 1.15, 0.9, 1.1]
    all_clips, clip_idx = [], 0
    for kw, folder in kw_to_folder.items():
        mp4s = list_all_mp4(folder)
        if not mp4s: continue
        n_pick = min(clips_per_keyword, len(mp4s))
        for i, p in enumerate(random.sample(mp4s, n_pick)):
            dur = random.uniform(CLIP_DURATION_MIN, CLIP_DURATION_MAX)
            speed = random.choice([0.85, 0.9, 1.0]) if i == 0 else speed_patterns[clip_idx % len(speed_patterns)]
            all_clips.append({"path": p, "source_start": random.uniform(0, 1.5), "duration": round(dur, 1), "keyword": kw, "speed": round(speed, 2)})
            clip_idx += 1
        print(f"   [DIR] {folder.name}: 选取 {n_pick}/{len(mp4s)} 个片段")
    return all_clips


def list_available_selling_points(min_files: int = 0) -> List[dict]:
    """列出所有可用的卖点文件夹"""
    base = Path(SELLING_POINT_DIR)
    if not base.exists(): return []
    result = []
    for d in sorted(base.iterdir()):
        if d.is_dir():
            count = len(list_all_mp4(d))
            if count < min_files: continue
            result.append({"folder": d, "name": re.sub(r'[\d+]', '', d.name).strip(), "original_name": d.name, "mp4_count": count})
    return result


# ═══════════════════════════════════════════════════════════════
# 文案-画面匹配
# ═══════════════════════════════════════════════════════════════

def _extract_visual_keywords(text: str) -> set:
    """从句子中提取可视化关键词"""
    style_words = ["意式中古风", "中古风", "包豪斯", "法式", "奶油风", "极简风",
                   "原木风", "轻奢风", "侘寂风", "工业风", "现代简约", "意式极简",
                   "悬浮式", "落地式", "开放式", "嵌入式", "复古"]
    material_words = ["岩板", "实木", "洞石", "奢石", "黑金", "潘多拉", "宝格丽",
                      "维多利亚", "柚木", "胡桃", "樱桃木", "橡木", "微水泥",
                      "大理石", "石英石", "纯黑", "白色", "奶白", "哑光", "亮光",
                      "普拉达绿", "法国金莱姆石", "香奈儿白", "罗马洞石"]
    function_words = ["薄抽", "抽屉", "烤箱", "蒸烤", "轨道插座", "插座", "水槽",
                      "灯带", "收纳", "伸缩", "餐桌", "酒柜", "冰箱", "电磁炉",
                      "烤炉", "拉篮", "吧台", "煮茶", "圆弧", "连纹", "钢结构",
                      "海棠角", "水磨边", "腰线", "亚克力腿", "岩板腿"]
    product_words = ["岛台", "餐边柜", "高柜", "岛头", "台面"]
    all_kw = style_words + material_words + function_words + product_words
    return {kw for kw in all_kw if kw in text}


def ai_match_clips(copy_text: str, num_clips: int = 8) -> list:
    """AI素材库智能匹配 v11.0 — FAISS向量优先 + 知识库规则补充, 无SQL降级"""
    # Layer 1: SmartMatcher (FAISS向量 + 知识库)
    if not _HAS_AI_LIBRARY or not _smart_matcher:
        raise RuntimeError(
            "SmartMatcher 不可用。向量检索是核心功能，禁止降级。\n"
            "请运行: python force_rebuild_faiss.py 重建FAISS索引\n"
            "或设置环境变量 TREECUT_SKIP_VECTOR_SEARCH=1 临时跳过（会降低匹配质量）"
        )

    try:
        results = _smart_matcher.search_by_copy(copy_text, num_clips)
        if results:
            clips = []
            for r in results[:num_clips * 2]:
                duration = float(r.get("end_time", 5)) - float(r.get("start_time", 0))
                if duration > 1.0:
                    clips.append({"path": Path(r["video_path"]), "source_start": float(r.get("start_time", 0)),
                                  "duration": min(duration, CLIP_DURATION_MAX), "speed": 1.0,
                                  "ai_matched": True, "match_score": r.get("score", 0),
                                  "match_method": r.get("match_method", "SmartMatcher")})
            if clips:
                print(f"   [SmartMatcher] 匹配: {len(clips)}个片段 (FAISS向量/知识库)")
                # CLIP 图文重排序 (v11.1 可选增强)
                try:
                    from core.clip_matcher import get_clip_matcher
                    cm = get_clip_matcher()
                    if cm and cm.available:
                        clips = cm.rerank_clips(copy_text, clips, top_k=num_clips)
                        print(f"   [CLIP] 图文重排序: {len(clips)}个片段")
                except Exception:
                    pass  # CLIP 不可用时静默跳过
                return clips[:num_clips]
    except Exception as e:
        print(f"   [ERROR] SmartMatcher 搜索失败: {e}")
        raise

    # Layer 2: 知识库规则 → SQL 查询数据库中的实际片段
    if _kb_bridge:
        kb_kws = _kb_bridge.extract_copy_keywords(copy_text)
        kb_found = sum(len(v) for v in kb_kws.values())
        if kb_found >= 2:
            print(f"   [Layer2] 知识库: 识别到{kb_found}个行业术语")
            kw_list = []
            for items in kb_kws.values():
                kw_list.extend(list(items)[:5])
            if kw_list:
                import sqlite3
                main_db = str(Path(__file__).parent / "ai_material_library.db")
                if Path(main_db).exists():
                    with sqlite3.connect(main_db) as conn:
                        # 修复：每个关键词单独匹配，用OR连接
                        # 原来的错误：用%连接多个关键词后做LIKE，几乎不可能命中
                        conditions = []
                        params = []
                        for kw in kw_list[:10]:
                            conditions.append("tags LIKE ?")
                            params.append(f"%{kw}%")
                            conditions.append("objects LIKE ?")
                            params.append(f"%{kw}%")
                        where_clause = " OR ".join(conditions)
                        sql = f"""
                            SELECT video_path, start_time, end_time, tags, objects, style, color, material
                            FROM materials 
                            WHERE {where_clause} 
                            LIMIT ?
                        """
                        params.append(num_clips * 2)
                        rows = conn.execute(sql, params).fetchall()
                    if rows:
                        # 计算每个记录的匹配关键词数量，按匹配数排序
                        scored_rows = []
                        for r in rows:
                            d = dict(zip(["video_path","start_time","end_time","tags","objects","style","color","material"], r))
                            match_count = 0
                            tags_lower = (d.get("tags") or "").lower()
                            objects_lower = (d.get("objects") or "").lower()
                            for kw in kw_list[:10]:
                                kw_lower = kw.lower()
                                if kw_lower in tags_lower:
                                    match_count += 1
                                if kw_lower in objects_lower:
                                    match_count += 1
                            scored_rows.append((match_count, d))
                        # 按匹配数降序排序
                        scored_rows.sort(key=lambda x: -x[0])
                        rows = [tuple(d.values()) for _, d in scored_rows]
                    if rows:
                        clips = []
                        for r in rows:
                            d = dict(zip(["video_path","start_time","end_time","tags","objects","style","color","material"], r))
                            duration = float(d.get("end_time", 5)) - float(d.get("start_time", 0))
                            if duration > 1.0:
                                clips.append({"path": Path(d["video_path"]), "source_start": float(d.get("start_time", 0)),
                                              "duration": min(duration, CLIP_DURATION_MAX), "speed": 1.0,
                                              "ai_matched": True, "match_score": 0.7,
                                              "match_method": "知识库规则"})
                        if clips:
                            print(f"   [Layer2] 知识库匹配: {len(clips)}个片段")
                            return clips[:num_clips]
    else:
        raise RuntimeError(
            "KnowledgeBridge 未初始化。知识库匹配是核心功能，禁止降级到纯文件名匹配。"
        )

    # 如果 FAISS 和知识库都返回空，提示重建索引
    print("   [WARN] FAISS 和知识库均未返回匹配结果。")
    print("   建议: python force_rebuild_faiss.py")
    return []


def _sort_by_freshness(clips: List[dict]) -> List[dict]:
    """按素材使用次数排序：使用少的优先 (v11.1 动态权重)"""
    try:
        from core.__init__ import MaterialUsageTracker
        return sorted(clips, key=lambda c: MaterialUsageTracker.get_usage_count(c["path"]))
    except Exception:
        return clips


def match_clips_to_sentences(subtitles: List[str], mp4_pool: List[Path],
                             folder_hints: dict = None) -> List[dict]:
    """
    逐句画面匹配 — 知识库增强版。
    用V5知识库提取关键词 → 文件名+文件夹名匹配 → 按匹配分排序选取。
    """
    used_paths = set()
    result = []

    for sentence in subtitles:
        # 1. 提取可视化关键词 (基础词汇 + 知识库词汇)
        vis_kws = _extract_visual_keywords(sentence)

        # 2. 用知识库补充更多关键词
        if _kb_bridge:
            try:
                kb_kws = _kb_bridge.extract_copy_keywords(sentence)
                for cat_items in kb_kws.values():
                    vis_kws.update(cat_items)
            except Exception as _e:
                from utils.logging import log_warning
                log_warning('pipeline', str(_e)[:80])

        # 3. 用关键词映射补充分类文件夹名
        folder_kws = set()
        for kw in list(vis_kws):
            if kw in KEYWORD_FOLDER_MAP:
                folder_kws.update(KEYWORD_FOLDER_MAP[kw])

        if not vis_kws or not mp4_pool:
            available = [m for m in mp4_pool if m not in used_paths]
            if not available:
                available = list(mp4_pool)
            pick = random.choice(available)
            dur = round(random.uniform(CLIP_DURATION_MIN, CLIP_DURATION_MAX), 1)
            result.append({"path": pick, "source_start": round(random.uniform(0, 3.0), 1),
                           "duration": dur, "speed": round(random.gauss(1.0, 0.04), 2),
                           "match_score": 0, "match_method": "random"})
            used_paths.add(pick)
            continue

        # 4. 对素材池中每个素材打分
        scored = []
        for mp4 in mp4_pool:
            if mp4 in used_paths:
                continue
            fname = mp4.name.lower()
            folder = mp4.parent.name.lower()
            searchable = fname + " " + folder

            score = 0
            for kw in vis_kws:
                kw_lower = kw.lower()
                if kw_lower in fname:
                    score += 15  # 文件名精确匹配 高分
                elif kw_lower in folder:
                    score += 8   # 文件夹名匹配 中分
                elif any(kw_lower in f for f in folder_kws if f.lower() in folder):
                    score += 5   # 关键词映射的文件夹 加分

            if score > 0:
                scored.append((score, mp4))

        scored.sort(key=lambda x: -x[0])

        if scored and scored[0][0] >= 15:
            # 高置信度匹配
            pick = scored[0][1]
            method = "keyword_precise"
        elif scored:
            # 中等置信度 — 从Top3随机
            top3 = scored[:min(3, len(scored))]
            pick = random.choice([x[1] for x in top3])
            method = "keyword_partial"
        else:
            # 无匹配 — 从其他卖点文件夹随机选(避免同文件夹重复)
            available = [m for m in mp4_pool if m not in used_paths]
            if not available:
                available = list(mp4_pool)
            pick = random.choice(available)
            method = "random_diverse"

        dur = round(random.uniform(CLIP_DURATION_MIN, CLIP_DURATION_MAX), 1)
        result.append({"path": pick, "source_start": round(random.uniform(0, 3.0), 1),
                       "duration": dur, "speed": round(random.gauss(1.0, 0.04), 2),
                       "match_score": scored[0][0] if scored else 0,
                       "match_method": method})
        used_paths.add(pick)

    # 统计匹配质量
    precise = sum(1 for r in result if r.get("match_method") == "keyword_precise")
    partial = sum(1 for r in result if r.get("match_method") == "keyword_partial")
    print(f"   📊 匹配质量: {precise}精确/{partial}部分/{len(result)-precise-partial}随机")

    return result


# ═══════════════════════════════════════════════════════════════
# 时间轴同步
# ═══════════════════════════════════════════════════════════════

def build_tts_synced_timeline(subtitles: List[str], tts_duration_sec: float,
                              video_clips: List[dict]) -> tuple:
    """
    字幕自然断句 — 仅在标点处断开，确保每句语义完整。
    说到哪里显示到哪里，最后一条覆盖到配音结束。
    """
    if tts_duration_sec <= 0 or not subtitles:
        return list(video_clips), []

    full_text = "".join(subtitles)
    total_chars = len(full_text)
    if total_chars == 0:
        return list(video_clips), []

    # Step 1: 按标点自然断句（不在词中间切断）
    phrases = []
    current = ""
    for ch in full_text:
        current += ch
        if ch in '。！？，、,.!?':
            if current.strip():
                phrases.append(current.strip())
            current = ""
    if current.strip():
        phrases.append(current.strip())

    # 合并过短的短语（<4字合并到下一句）
    merged = []
    i = 0
    while i < len(phrases):
        current = phrases[i]
        # 如果当前太短且不是最后一条，合并下一条
        while len(current) < 4 and i + 1 < len(phrases):
            i += 1
            current += phrases[i]
        # 如果仍然太短但已是最后一条，保留
        if len(current) >= 2:
            merged.append(current)
        i += 1

    # 拆分过长短语（>18字在逗号处再拆一次）
    final_phrases = []
    for p in merged:
        if len(p) <= 18:
            final_phrases.append(p)
        else:
            # 在逗号处拆分
            sub_parts = re.split(r'(?<=[，,、])', p)
            buf = ""
            for sp in sub_parts:
                if len(buf) + len(sp) <= 16:
                    buf += sp
                else:
                    if buf:
                        final_phrases.append(buf)
                    buf = sp
            if buf:
                final_phrases.append(buf)

    if not final_phrases:
        final_phrases = [full_text]

    # Step 2: 按字符数比例精确分配时长 (零间隙，背靠背)
    chars_per_sec = total_chars / tts_duration_sec
    subtitle_timings = []
    current_sec = 0.0

    for i, phrase in enumerate(final_phrases):
        dur = len(phrase) / chars_per_sec
        dur = max(0.4, min(5.0, dur))
        is_last = (i == len(final_phrases) - 1)

        # 最后一条精确延伸到配音结束
        if is_last:
            dur = tts_duration_sec - current_sec
            dur = max(0.3, dur)
        elif current_sec + dur > tts_duration_sec:
            dur = max(0.3, tts_duration_sec - current_sec)

        if dur > 0.2:
            subtitle_timings.append({
                "text": phrase,
                "start_sec": round(current_sec, 3),
                "duration_sec": round(dur, 3),
            })

        current_sec += dur  # 零间隙！下一条紧接当前条结束
        if current_sec >= tts_duration_sec:
            break

    return list(video_clips), subtitle_timings


# ═══════════════════════════════════════════════════════════════
# 验证
# ═══════════════════════════════════════════════════════════════

def validate_draft(clips: List[dict], copy_text: str, tts_path, draft_dir: Path) -> tuple:
    """生成后自动复盘验证"""
    issues = []
    if ENABLE_PERSON_FILTER:
        for i, clip in enumerate(clips):
            fname = str(clip["path"]).lower()
            if any(kw in fname for kw in ["人像", "人物", "口播", "采访", "自拍", "露脸", "真人", "主播"]):
                issues.append(f"片段{i+1}含人像标记(未过滤): {clip['path'].name[:40]}")
    sentences = [s for s in re.split(r'[.。！？]', copy_text) if len(s) > 4]
    for i in range(len(sentences) - 2):
        if sentences[i][:6] == sentences[i+1][:6]:
            issues.append(f"文案重复: '{sentences[i][:20]}...'")
    if len(copy_text) < 60:
        issues.append(f"文案过短({len(copy_text)}字)")
    return len(issues) == 0, issues


# ═══════════════════════════════════════════════════════════════
# Excel 脚本读取
# ═══════════════════════════════════════════════════════════════

def read_script_excel(file_path: Path) -> Optional[dict]:
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
        copy_col = None
        for i, h in enumerate(headers):
            if any(kw in h.lower() for kw in ["文案", "脚本", "配音", "口播", "内容", "字幕"]):
                copy_col = i; break
        if copy_col is None:
            wb.close(); return None
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[copy_col] and str(row[copy_col]).strip():
                rows.append({"row_num": len(rows) + 2, "copy": str(row[copy_col]).strip(), "material_keyword": ""})
        wb.close()
        return {"file": file_path, "rows": rows, "copy_column": headers[copy_col], "total_rows": len(rows)}
    except Exception:
        return None


def get_script_row(script_path: str, row_num: int = 2) -> Optional[dict]:
    result = read_script_excel(Path(script_path))
    if not result or not result["rows"]: return None
    idx = max(0, min(row_num - 2, len(result["rows"]) - 1))
    row = result["rows"][idx]
    return {"copy": row["copy"], "material_keyword": row["material_keyword"], "source": f"{Path(script_path).name} 第{row['row_num']}行"}


# ═══════════════════════════════════════════════════════════════
# BGM 主题检测
# ═══════════════════════════════════════════════════════════════

def detect_video_theme(keyword: str, use_effects: bool = False) -> str:
    kw_lower = keyword.lower()
    factory_keywords = ["钢结构", "物流", "折边", "工艺", "材质", "细节", "生产", "工厂", "车间", "制作", "加工", "水磨边", "海棠角", "连纹", "圆弧"]
    for fk in factory_keywords:
        if fk in keyword or fk in kw_lower: return "工厂实力"
    effect_keywords = ["封面", "尺寸", "颜色", "造型", "生活化", "响指", "转场", "餐桌", "空间", "设计", "齐屏", "吧台"]
    for ek in effect_keywords:
        if ek in keyword or ek in kw_lower: return "效果展示"
    return "卖点展示"


# ═══════════════════════════════════════════════════════════════
# BGM 收集（简化版）
# ═══════════════════════════════════════════════════════════════

def collect_bgm(base_dir: str = None) -> List[Path]:
    if base_dir is None: base_dir = BGM_DIR
    base = Path(base_dir)
    results = []
    if base.exists():
        for ext in ["*.mp3", "*.wav", "*.flac", "*.aac", "*.m4a", "*.ogg"]:
            results.extend(base.rglob(ext))
    if not results and base_dir == BGM_PATH_PRIMARY:
        fallback = Path(BGM_DIR_FALLBACK)
        if fallback.exists():
            for ext in ["*.mp3", "*.wav", "*.flac", "*.aac", "*.m4a", "*.ogg"]:
                results.extend(fallback.rglob(ext))
    seen = {}
    for f in sorted(results, key=lambda x: x.name):
        key = (f.name, f.stat().st_size)
        if key not in seen: seen[key] = f
    return list(seen.values())


# ═══════════════════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════════════════

def print_banner():
    print(r"""
╔══════════════════════════════════════════════╗
║       🌳 树剪 · AI视频自动剪辑工具            ║
║           剪映专业版草稿自动生成              ║
╚══════════════════════════════════════════════╝""")


def print_clip_table(clips: List[dict]):
    print(f"\n   [*] 选取了 {len(clips)} 个片段:")
    total_dur = 0
    for i, clip in enumerate(clips):
        dur = clip["duration"]; total_dur += dur
        fname = clip["path"].name
        if len(fname) > 55: fname = fname[:52] + "..."
        print(f"   [{i+1}] {fname}")
        print(f"        ↳ 截取 {clip['source_start']:.1f}s 起 {dur:.1f}秒")
    print(f"   [*] 总时长: {total_dur:.1f}秒")
    return total_dur


def save_copy(copy_text: str, keyword: str):
    Path(OUTPUT_COPY_DIR).mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = re.sub(r'[^\w一-鿿]', '_', keyword)
    filepath = Path(OUTPUT_COPY_DIR) / f"文案_{safe_keyword}_{timestamp}.txt"
    filepath.write_text(copy_text, encoding="utf-8")
    print(f"   [*] 文案已保存: {filepath}")
    return filepath


def run(keyword: str, num_clips: int = None, bgm_file: str = None,
        use_effects: bool = False, dry_run: bool = False,
        auto_bgm: bool = False, generate_tts: bool = False,
        auto_mode: bool = False,
        tts_voice: str = None, custom_voice: bool = False,
        use_script: str = None, script_row: int = 2,
        direct_copy: str = None, copy_text_override: str = None,
        progress_callback=None, script_id: int = None,
        exclude_paths: list = None, task_id: str = None,
        video_log_id: int = None, video_index: int = 1,
        session_used_paths: set = None, script_hash: str = None,
        no_script_library: bool = False):
    """主生成管道。
       exclude_paths: 会话级已用素材路径列表。
       session_used_paths: 本次批量任务已用素材集合(set) — v11.6强制去重。
       script_hash: 当前脚本哈希 — v11.6脚本偏好学习。
       task_id: 生成任务ID — 用于持久化素材使用记录。
       video_log_id: 关联 generation_video_log 的ID。
       video_index: 批量中的序号。
       no_script_library: 用户明确要求不使用脚本库时设为True。默认False=优先从脚本库随机选取。
    """
    TOTAL_STEPS = 6

    def _progress(step, msg):
        if progress_callback:
            progress_callback(step, TOTAL_STEPS, msg)

    print_banner()
    print(f">> 卖点关键词: 「{keyword}」")

    if num_clips is None: num_clips = NUM_CLIPS_MAX

    # Step 1: 多卖点混合素材扫描
    _progress(1, "素材扫描")
    print(f"\n{'='*50}")
    print(f"[SCAN] Step 1/6: 多卖点混合素材扫描...")
    multi = collect_multi_point_mp4s(keyword)
    primary_mp4s = multi["primary"]; primary_name = multi["primary_name"]
    other_points = multi["other_points"]; global_mp4s = multi["global_show"]
    all_other_mp4s = multi["all_other_mp4s"]
    b_group_mp4s = collect_b_group_mp4s(keyword) if ENABLE_B_GROUP_MIX else []
    all_mp4s = primary_mp4s + all_other_mp4s + global_mp4s + b_group_mp4s
    if not all_mp4s:
        print(f"\n[FAIL] 所有素材池为空！")
        return {
            "success": False,
            "error": "所有素材池为空，请检查素材目录配置",
            "keyword": keyword,
            "clips": [],
            "copy_text": "",
            "draft_dir": None,
            "bgm_file": bgm_file,
        }

    # Step 2: 获取文案
    print(f"\n{'='*50}")
    script_copy = None; script_keyword = None
    if copy_text_override:
        script_copy = copy_text_override
        _progress(2, "AI文案生成"); print(f"[*] Step 2/6: 使用外部文案({len(script_copy)}字)...")
    if not script_copy and direct_copy:
        script_copy = direct_copy
        _progress(2, "AI文案生成"); print(f"[*] Step 2/6: 使用指定文案({len(script_copy)}字)...")
    if not script_copy and use_script:
        script_path = Path(SCRIPT_FOLDER_PATH) / use_script
        if not script_path.exists(): script_path = Path(use_script)
        if script_path.exists():
            row_data = get_script_row(str(script_path), script_row)
            if row_data and row_data["copy"]:
                script_copy = row_data["copy"]
                _progress(2, "AI文案生成"); print(f"[*] Step 2/6: 使用脚本文案({row_data['source']})...")

    # ── v11.8 硬性规则: 默认从脚本库随机选取，除非用户明确说不用 ──
    if not script_copy and not no_script_library:
        try:
            from core.script_learning import get_library
            lib = get_library()
            chosen = lib.get_script_with_diversity(prefer_unused=True)
            if chosen:
                script_copy = chosen["content"]
                script_id = chosen["id"]
                _progress(2, "AI文案生成"); print(f"[*] Step 2/6: 📜 从脚本库随机选取 (ID={chosen['id']}, 已用{chosen['usage_count']}次)...")
                print(f"   [ScriptLib] 选中脚本预览: {script_copy[:80]}...")
            else:
                print(f"   [ScriptLib] ⚠️ 脚本库为空, 降级到AI生成")
        except Exception as e:
            print(f"   [ScriptLib] 脚本库查询失败(非致命): {e}, 降级到AI生成")

    if script_copy:
        copy_text = script_copy
    else:
        copy_text = generate_copy(keyword, num_clips, DEFAULT_COPY_DURATION, clips=None)

    print(f"\n{'─'*40}\n{copy_text}\n{'─'*40}")
    copy_path = save_copy(copy_text, keyword)

    # 字幕拆分
    subtitles = split_copy_to_subtitles(copy_text)
    subtitles = [strip_leading_junk(s) for s in subtitles]
    subtitles = [s for s in subtitles if s and len(re.findall(r'[一-鿿]', s)) >= 2]
    if not subtitles: subtitles = ["坤宝岛台"]
    
    # 修复：连接前先去掉每个字幕末尾的标点，避免重复标点
    cleaned_subtitles = [re.sub(r'[.。！？!?，,；;]+$', '', s) for s in subtitles]
    tts_text = "。".join(cleaned_subtitles)
    tts_text = clean_text_for_tts(tts_text)
    est_tts = estimate_tts_duration(tts_text)
    print(f"   [*] {len(subtitles)}条字幕(句子级拆分,已清洗)")

    # Step 3: 配音
    print(f"\n{'='*50}")
    _progress(3, "AI配音"); print(f"[*]  Step 3/6: 生成完整配音...")
    tts_path = None; tts_duration = 0.0
    if generate_tts:
        actual_voice = tts_voice or TTS_VOICE
        tts_path = generate_tts_voiceover(tts_text, keyword, voice=actual_voice)
        if tts_path:
            tts_duration = get_audio_duration_seconds(tts_path)
            print(f"   ⏱  ★ 配音精确时长: {tts_duration:.1f}秒")
        else:
            print(f"   !! 配音生成失败,使用估算时长")
            tts_duration = est_tts
    else:
        print(f"   [TIP] 使用 --tts 参数可自动生成 AI 配音")
        tts_duration = est_tts

    target_video_dur = max(tts_duration, TARGET_DURATION_MIN) if tts_duration > 0 else DEFAULT_COPY_DURATION

    # Step 4: 根据配音时长选取素材
    print(f"\n{'='*50}")
    _progress(4, "素材匹配"); print(f"[*] Step 4/6: 根据配音时长选取素材(目标{target_video_dur:.1f}秒)...")
    # v11.6: 自适应匹配 (三层去重+偏好学习+热度衰减)
    # ── v11.7: 开头素材优先从 is_opening 池中选取 ──
    opening_clip = None
    if subtitles and len(subtitles) > 1:
        try:
            from core.database import db
            with db.get_connection() as conn:
                exclude_cond = ""
                params = []
                if exclude_paths:
                    placeholders = ",".join(["?"] * len(exclude_paths))
                    exclude_cond = f"AND video_path NOT IN ({placeholders})"
                    params = list(exclude_paths)
                rows = conn.execute(f"""
                    SELECT video_path, start_time, end_time, duration
                    FROM materials
                    WHERE is_opening = 1 AND has_human = 0 AND analyzed = 1 {exclude_cond}
                    ORDER BY RANDOM() LIMIT 1
                """, params).fetchall()
                if not rows:
                    rows = conn.execute(f"""
                        SELECT video_path, start_time, end_time, duration
                        FROM materials
                        WHERE has_human = 0 AND analyzed = 1 {exclude_cond}
                        ORDER BY RANDOM() LIMIT 1
                    """, params).fetchall()
                if rows:
                    r = rows[0]
                    opening_clip = {
                        "path": Path(r[0]),
                        "source_start": float(r[1]) if r[1] else 0,
                        "duration": min(float(r[3]) if r[3] and r[3] > 0 else CLIP_DURATION_MAX, CLIP_DURATION_MAX),
                        "speed": 1.0,
                        "match_method": "opening_pool",
                        "match_score": 1.0,
                    }
                    print(f"   [Opening] 开场素材: {Path(r[0]).name[:40]}")
        except Exception as e:
            print(f"   [Opening] 开场素材查询失败(非致命): {e}")

    session_used = session_used_paths or set()

    # 把开头素材加入排除集，后续匹配不再选它
    if opening_clip:
        op_path = str(opening_clip["path"])
        session_used.add(op_path)
        if exclude_paths:
            exclude_paths = list(exclude_paths) + [op_path]
    if exclude_paths:
        session_used = session_used.union(set(exclude_paths))
    shash = script_hash or (make_script_hash(copy_text) if copy_text else None)

    # 如有开场专用素材，剩余句子减少1条
    remaining_subs = len(subtitles) - (1 if opening_clip else 0)
    if shash:
        print(f"   [Adaptive] 脚本哈希: {shash}, 已排除: {len(session_used)}个素材")
        candidates = get_adaptive_clips(copy_text, max(1, remaining_subs), session_used, shash)
        if candidates:
            ai_clips = convert_to_clips(candidates, len(subtitles))
            print(f"   [AdaptiveMatch] 自适应匹配: {len(ai_clips)}个片段 (去重+偏好+衰减)")
        else:
            ai_clips = ai_match_clips(copy_text, len(subtitles))
    else:
        ai_clips = ai_match_clips(copy_text, len(subtitles))

    # v11.1: 脚本语义理解增强 (可选 — 需 DeepSeek API Key)
    if not ai_clips and DEEPSEEK_API_KEY:
        try:
            from core.script_understanding import ScriptParser
            parser = ScriptParser(use_ai=True)
            parsed = parser.parse(copy_text)
            if parsed.get("segments") and _smart_matcher:
                raw_results = _smart_matcher.search_by_visual_requirements(parsed, len(subtitles))
                ai_clips = convert_to_clips(raw_results, len(subtitles))
                print(f"   [ScriptUnderstanding] 语义匹配: {len(ai_clips)}个片段")
        except Exception as e:
            print(f"   [ScriptUnderstanding] 跳过 (需要DeepSeek API Key): {e}")

    if opening_clip:
        if ai_clips:
            clips = [opening_clip] + ai_clips
        else:
            clips = [opening_clip] + match_clips_to_sentences(
                subtitles[1:], all_mp4s) if len(subtitles) > 1 else [opening_clip]
    elif ai_clips:
        clips = ai_clips
    else:
        clips = match_clips_to_sentences(subtitles, all_mp4s)

    # v11.1: 动态素材权重 — 使用少的优先
    clips = _sort_by_freshness(clips)

    # ── 人像过滤 (v11.5: 强制产品优先，绝不容忍人像) ──
    human_kw = ["人像","人物","口播","主播","真人","采访","露脸","自拍",
                "person","face","portrait","people","human"]
    
    def _has_human_content(clip):
        """检查素材是否包含人像内容"""
        # 检查has_human标记
        if isinstance(clip.get("has_human"), int) and clip.get("has_human", 0) == 1:
            return True
        # 检查路径、标签、对象中是否有人像关键词
        search_fields = [
            str(clip.get("path", "")).lower(),
            str(clip.get("tags", "")).lower(),
            str(clip.get("objects", "")).lower(),
        ]
        for kw in human_kw:
            kw_lower = kw.lower()
            for field in search_fields:
                if kw_lower in field:
                    return True
        return False
    
    before_h = len(clips)
    clips = [c for c in clips if not _has_human_content(c)]
    if len(clips) < before_h:
        print(f"   [FILTER] 人像过滤: {before_h} → {len(clips)} 个片段 (强制产品优先)")
    # 绝不用人像补充 — 过滤后不足则用纯文件名匹配的随机素材
    if len(clips) < 3 and all_mp4s:
        fallback = [m for m in all_mp4s if not any(kw in str(m).lower() for kw in human_kw)]
        random.shuffle(fallback)
        needed = max(3 - len(clips), 2)
        for m in fallback[:needed]:
            clips.append({"path": m, "source_start": round(random.uniform(0, 3.0), 1),
                          "duration": round(random.uniform(CLIP_DURATION_MIN, CLIP_DURATION_MAX), 1),
                          "speed": 1.0, "match_score": 0, "match_method": "产品补充(无人像)"})
        print(f"   [!] 强制产品补充 {min(needed, len(fallback))} 个片段(已排除人像)")

    # ── 排除已用素材 (v11.2 批量去重) ──
    if exclude_paths:
        before_e = len(clips)
        exclude_set = {str(p) for p in exclude_paths}
        clips = [c for c in clips if str(c["path"]) not in exclude_set]
        if len(clips) < before_e:
            print(f"   [DEDUP] 批量去重: {before_e} → {len(clips)} 个片段 (排除{len(exclude_set)}个已用)")
    total_dur = print_clip_table(clips)

    # ── 质量过滤 (v10.3) ──
    if ENABLE_QUALITY_FILTER:
        from utils.quality_scorer import get_scorer
        scorer = get_scorer()
        before = len(clips)
        clips = scorer.filter_by_quality(clips, min_score=QUALITY_MIN_SCORE)
        print(f"   >> 质量过滤: {before} → {len(clips)} 个片段 (阈值{QUALITY_MIN_SCORE})")
        if ENABLE_DEDUPLICATION:
            before2 = len(clips)
            clips = scorer.deduplicate(clips, hamming_distance=DEDUPLICATION_HAMMING_DIST)
            print(f"   [DEDUP] 去重: {before2} → {len(clips)} 个片段")
        # 过滤后不足则降低阈值补充
        if len(clips) < 3:
            backup = match_clips_to_sentences(subtitles, all_mp4s)
            backup = [c for c in backup if c not in clips]
            needed = max(3, len(clips)) - len(clips)
            clips.extend(backup[:needed])
            print(f"   [!] 过滤过强, 补充 {needed} 个片段")
        # 重新计算总时长
        total_dur = sum(c.get("duration",CLIP_DURATION_MIN) for c in clips)

    # 素材不足时自动补充
    MIN_BUFFER = 1.5
    if tts_duration > 0 and total_dur < tts_duration + MIN_BUFFER:
        shortage = (tts_duration + MIN_BUFFER) - total_dur
        print(f"   [!] 素材不足: 缺{shortage:.1f}s,自动补充...")
        used_paths = {c["path"] for c in clips}
        pool = [m for m in all_mp4s if m not in used_paths]
        random.shuffle(pool)
        added = 0
        while total_dur < tts_duration + MIN_BUFFER and pool and added < 6:
            extra = pool.pop()
            dur = min(CLIP_DURATION_MAX, (tts_duration + MIN_BUFFER) - total_dur + 0.5)
            dur = round(random.uniform(CLIP_DURATION_MIN, max(CLIP_DURATION_MIN, dur)), 1)
            clips.append({"path": extra, "source_start": round(random.uniform(0, 3.0), 1), "duration": dur, "speed": round(random.gauss(1.0, 0.04), 2)})
            total_dur += dur; added += 1
        if added > 0:
            print(f"   [OK] 自动补充 {added} 个片段 -> 总时长{total_dur:.1f}s")
            print_clip_table(clips)

    # Step 5: BGM
    print(f"\n{'='*50}")
    _progress(5, "BGM处理"); print(f"[*] Step 5/6: 处理BGM...")
    bgm_path = None
    if bgm_file:
        bgm_path = Path(bgm_file)
    if bgm_path is None and auto_bgm:
        bgm_path = _auto_get_bgm_simple(keyword, target_video_dur)
    if bgm_path is None:
        bgms = collect_bgm()
        if bgms: bgm_path = random.choice(bgms)

    # Step 6: 生成剪映草稿
    print(f"\n{'='*50}")
    _progress(6, "草稿输出"); print(f"[*] Step 6/6: 生成剪映草稿...")

    if dry_run:
        print(f"\n   [SCAN] [DRY RUN] 不会实际生成文件")
        return {"clips": clips, "copy": copy_text, "subtitles": subtitles}

    if tts_duration > 0:
        adj_clips, subtitle_timings = build_tts_synced_timeline(subtitles, tts_duration, clips)
        clips_to_use = adj_clips
    else:
        clips_to_use = clips
        subtitle_timings = []
        current_t = 0.0
        for i, sub in enumerate(subtitles):
            d = clips[i].get("duration", 2.5) if i < len(clips) else 2.5
            subtitle_timings.append({"text": sub, "start_sec": current_t, "duration_sec": d})
            current_t += d

    try:
        builder = JianyingDraftBuilder(draft_name=f"树剪_{keyword}", keyword=keyword)
        for clip in clips_to_use:
            builder.add_video_clip(clip["path"], clip["source_start"], clip["duration"], clip.get("speed", 1.0))
        if tts_path and ENABLE_AUTO_TRIM_VIDEO and tts_duration > 0:
            builder.trim_all_to_tts(tts_duration)
        effective_duration = builder._total_duration_us
        if bgm_path:
            builder.add_audio_bgm(bgm_path)
        if tts_path:
            builder.add_tts_audio(tts_path)
        last_end_us = 0
        total_subs = len(subtitle_timings)
        for idx, timing in enumerate(subtitle_timings):
            start_us = sec_to_us(timing["start_sec"])
            dur_us = sec_to_us(timing["duration_sec"])
            if start_us < last_end_us: start_us = last_end_us
            if dur_us <= 100_000: continue
            is_last = (idx == total_subs - 1)
            if is_last: dur_us = effective_duration - start_us
            elif start_us + dur_us > effective_duration: dur_us = effective_duration - start_us
            if dur_us > 200_000:
                builder.add_subtitle(timing["text"], start_us, dur_us)
                last_end_us = start_us + dur_us
            if is_last: break
        draft_dir = save_draft(builder, keyword)
        # 保存视频元数据 (v11.2 反馈学习)
        if video_log_id and draft_dir:
            try:
                metadata = {"video_log_id": video_log_id, "keyword": keyword,
                            "total_duration": total_dur, "tts_duration": tts_duration}
                (draft_dir / "video_metadata.json").write_text(
                    json.dumps(metadata, ensure_ascii=False), encoding="utf-8")
            except Exception: pass
    except ImportError as e:
        print(f"   [FAIL] 无法生成草稿: {e}")
        print(f"   [TIP] 请安装: pip install pyJianYingDraft")
        return {"error": str(e)}

    # 验证
    is_valid, issues = validate_draft(clips_to_use, copy_text, tts_path, draft_dir)
    if not is_valid:
        print(f"\n   !! 复盘发现 {len(issues)} 个问题(仅提示):")
        for issue in issues: print(f"      - {issue}")

    print(f"\n{'='*50}")
    print(f"[OK] 全部完成！")
    print(f"{'='*50}")

    # v11.6: 持久化素材使用记录 + 视频日志 + 脚本偏好学习
    if task_id:
        try:
            from core.database import db
            db.record_material_usage(task_id, clips_to_use)
            if video_log_id is None:
                video_log_id = db.insert_video_log(
                    task_id, video_index, keyword, copy_text, total_dur, tts_duration)
            for idx, clip in enumerate(clips_to_use):
                db.insert_material_log(video_log_id, clip, idx)
                # 脚本偏好学习
                if shash:
                    record_script_match(shash, str(clip.get("path", "")))
            print(f"   [MatLog] video#{video_log_id}: {len(clips_to_use)} clips recorded" +                   (f", 偏好已记录(shash={shash})" if shash else ""))
        except Exception as e:
            print(f"   [MatLog] 记录失败: {e}")

    # 脚本学习库: 如果提供了 script_id，更新使用统计
    if script_id:
        try:
            from core.script_learning import get_library
            get_library().update_score(script_id, 3.0)
            print(f"   [ScriptLib] 脚本 #{script_id} 使用统计已更新")
        except Exception as e:
            print(f"   [ScriptLib] 更新失败: {e}")

    return {"draft_dir": str(draft_dir), "clips": clips, "copy": copy_text,
            "subtitles": subtitles, "total_duration": total_dur,
            "tts_duration": tts_duration, "tts_path": str(tts_path) if tts_path else None,
            "video_log_id": video_log_id}


def _auto_get_bgm_simple(keyword: str, duration_sec: float) -> Optional[Path]:
    """
    自动BGM获取 — 本地文件随机选择，按视频主题匹配风格。
    """
    bgms = collect_bgm()
    if bgms:
        theme = detect_video_theme(keyword)
        mood_map = {"卖点展示": "upbeat", "效果展示": "chill", "工厂实力": "cinematic"}
        mood = mood_map.get(theme, "ambient")
        mood_bgms = [b for b in bgms if mood in b.stem.lower()]
        chosen = random.choice(mood_bgms if mood_bgms else bgms)
        print(f"   [*] [BGM] 本地随机: {chosen.name}")
        return chosen

    print(f"   [TIP] [BGM] 未找到本地BGM，请在 {BGM_PATH_PRIMARY} 放入音乐文件")
    return None


def run_multi(keywords: List[str], clips_per_kw: int = 2, bgm_file: str = None,
              auto_bgm: bool = False, dry_run: bool = False, generate_tts: bool = False,
              tts_voice: str = None, template: str = "multi", use_effects: bool = True,
              custom_voice: bool = False, no_script_library: bool = False):
    """多卖点混剪主流程（简化版，核心逻辑保持一致）"""
    print_banner()
    kw_display = ",".join(keywords)
    print(f">> 多卖点混剪模式: 「{kw_display}」")

    clips = collect_multi_selling_mp4s(keywords, clips_per_kw)
    if not clips:
        print(f"\n[FAIL] 没有找到任何匹配的素材！")
        return None

    total_dur = sum(c["duration"] for c in clips)
    print(f"\n   [*] 共选取 {len(clips)} 个片段,总时长 {total_dur:.1f}秒")

    kw_set = set(c["keyword"] for c in clips)
    # ── v11.8 默认从脚本库随机选取 ──
    library_copy = None; script_id = None
    if not no_script_library:
        try:
            from core.script_learning import get_library
            lib = get_library()
            chosen = lib.get_script_with_diversity(prefer_unused=True)
            if chosen:
                library_copy = chosen["content"]
                script_id = chosen["id"]
                print(f"\n   [ScriptLib] 📜 从脚本库选取 (ID={chosen['id']}, 已用{chosen['usage_count']}次)")
        except Exception as e:
            print(f"   [ScriptLib] 脚本库不可用: {e}")
    copy_text = library_copy or generate_fallback_copy("+".join(sorted(kw_set)))
    print(f"\n   [*] 文案:\n{'─'*40}\n{copy_text}\n{'─'*40}")

    keyword_str = "+".join(sorted(kw_set))[:40]
    copy_path = save_copy(copy_text, keyword_str)

    subtitles = split_copy_to_subtitles(copy_text)
    subtitles = [strip_leading_junk(s) for s in subtitles]
    subtitles = [s for s in subtitles if s and len(re.findall(r'[一-鿿]', s)) >= 2]
    if not subtitles: subtitles = ["坤宝岛台"]
    
    # 修复：连接前先去掉每个字幕末尾的标点，避免重复标点
    cleaned_subtitles = [re.sub(r'[.。！？!?，,；;]+$', '', s) for s in subtitles]
    tts_text = "。".join(cleaned_subtitles)
    tts_text = clean_text_for_tts(tts_text)

    tts_path = None; tts_duration = 0.0
    if generate_tts:
        actual_voice = tts_voice or TTS_VOICE
        tts_path = generate_tts_voiceover(tts_text, keyword_str, voice=actual_voice)
        if tts_path: tts_duration = get_audio_duration_seconds(tts_path)

    bgm_path = None
    if bgm_file: bgm_path = Path(bgm_file)
    if bgm_path is None and auto_bgm:
        bgm_path = _auto_get_bgm_simple(keyword_str, total_dur)
    if bgm_path is None:
        bgms = collect_bgm()
        if bgms: bgm_path = random.choice(bgms)

    if dry_run:
        print(f"\n   [SCAN] [DRY RUN] 不会实际生成文件")
        return {"clips": clips, "copy": copy_text, "subtitles": subtitles}

    if tts_duration > 0:
        adj_clips, subtitle_timings = build_tts_synced_timeline(subtitles, tts_duration, clips)
        clips_to_use = adj_clips
    else:
        clips_to_use = clips
        subtitle_timings = []
        current_t = 0.0
        for i, sub in enumerate(subtitles):
            d = clips[i].get("duration", 3.0) if i < len(clips) else 2.0
            subtitle_timings.append({"text": sub, "start_sec": current_t, "duration_sec": d})
            current_t += d

    try:
        builder = JianyingDraftBuilder(draft_name=f"树剪混剪_{keyword_str[:30]}", keyword=keyword_str)
        for clip in clips_to_use:
            builder.add_video_clip(clip["path"], clip["source_start"], clip["duration"], clip.get("speed", 1.0))
        if tts_path and ENABLE_AUTO_TRIM_VIDEO and tts_duration > 0:
            builder.trim_all_to_tts(tts_duration)
        effective_duration = builder._total_duration_us
        if bgm_path: builder.add_audio_bgm(bgm_path)
        if tts_path: builder.add_tts_audio(tts_path)
        last_end_us = 0
        for timing in subtitle_timings:
            start_us = sec_to_us(timing["start_sec"]); dur_us = sec_to_us(timing["duration_sec"])
            if start_us < last_end_us: start_us = last_end_us
            if dur_us <= 0: continue
            if start_us + dur_us > effective_duration: dur_us = effective_duration - start_us
            if dur_us > 0:
                builder.add_subtitle(timing["text"], start_us, dur_us)
                last_end_us = start_us + dur_us
        draft_dir = save_draft(builder, keyword_str)
    except ImportError as e:
        print(f"   [FAIL] 无法生成草稿: {e}")
        return {"error": str(e)}

    print(f"\n{'='*50}")
    print(f"[OK] 全部完成！")
    return {"draft_dir": str(draft_dir), "clips": clips, "copy": copy_text,
            "subtitles": subtitles, "total_duration": total_dur}


def run_batch(keyword, count, script_path=None, interval_sec=3.0,
             no_script_library: bool = False, **kwargs):
    """批量生产 (v11.2: 会话去重 + 持久化历史去重 + 内存监控 + GC)
       no_script_library: 用户明确要求不用脚本库时设为True"""
    # 传入 kwargs 以便 run() 接收 no_script_library
    if 'no_script_library' not in kwargs:
        kwargs['no_script_library'] = no_script_library
    results = []; success = 0; fail = 0
    used_paths = []  # 会话级已用素材
    # 加载持久化历史 — 最近30天内所有已用素材
    try:
        from core.database import db
        history_paths = db.get_used_material_paths(days=30)
        used_paths.extend(history_paths)
        print(f"  [Batch] 加载 {len(history_paths)} 条历史已用素材 (30天)")
    except Exception:
        pass
    # 生成本次任务ID
    import random
    task_id = f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{random.randint(1000,9999)}"
    print(f"\n{'='*50}")
    print(f"  [Batch] Task: {task_id}")
    print(f"  [Batch] Generating {count} videos (dedup ON, {len(used_paths)} excluded)")
    for i in range(1, count + 1):
        print(f"\n  >> [{i}/{count}] Starting...")
        try:
            copy_text = None
            if script_path:
                row_data = get_script_row(script_path, row_num=i + 1)
                copy_text = row_data.get("copy", row_data.get("copy_text", "")) if row_data else ""
                if not copy_text:
                    print(f"  [Skip] Script row {i+1} is empty"); continue
            result = run(keyword=keyword, copy_text_override=copy_text,
                         exclude_paths=list(used_paths), task_id=task_id,
                         session_used_paths=set(used_paths),
                         script_hash=make_script_hash(copy_text) if copy_text else None,
                         **kwargs)
            if result and "clips" in result:
                # v11.6: 开头素材优先加入黑名单，确保下次不同
                clips = result["clips"]
                if clips:
                    opening = str(clips[0].get("path", ""))
                    if opening and opening not in used_paths:
                        used_paths.insert(0, opening)
                for c in clips:
                    p = str(c.get("path", ""))
                    if p and p not in used_paths:
                        used_paths.append(p)
            success += 1
            results.append({"index": i, "status": "success", "result": result})
            print(f"  [OK] [{i}/{count}] Done")
        except Exception as e:
            fail += 1
            results.append({"index": i, "status": "failed", "error": str(e)})
            print(f"  [FAIL] [{i}/{count}]: {e}")
            traceback.print_exc()
        # ── v11.2: 内存监控 + 资源释放 ──
        try:
            import psutil
            mem = psutil.virtual_memory()
            if mem.percent > 80:
                print(f"   [MEM] 内存使用 {mem.percent:.0f}%, 暂停5秒...")
                time.sleep(5)
        except ImportError:
            pass  # psutil 未安装，跳过
        gc.collect()
        try:
            import torch
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
        except Exception:
            pass
        if i < count: time.sleep(interval_sec)
    print(f"\n{'='*50}")
    print(f"  [Batch Done] Success: {success} / Failed: {fail} / Total: {count}")
    return results


# ═══════════════════════════════════════════════════════════════
# CLI 入口
# ═══════════════════════════════════════════════════════════════

# CLI 入口统一在 树剪.py — 此模块仅提供 run/run_multi/run_batch 函数
if __name__ == "__main__":
    print("请使用: python 树剪.py [关键词] [选项]")
    print("或:     python 树剪.py --web")
    print("或:     python 树剪.py --setup")
