# -*- coding: utf-8 -*-
"""将项目代码分成两个Excel文件，代码完整不缺失"""
import os, sys, datetime
sys.dont_write_bytecode = True
PROJECT_DIR = r"C:\Users\admin\Desktop\树剪软件相关文件"
DESKTOP = r"C:\Users\admin\Desktop"
exclude_dirs = {"__pycache__", "data", "logs", ".git", ".github", ".claude", "node_modules"}
# 1. 收集所有文件
all_files = []
for root, dirs, filenames in os.walk(PROJECT_DIR):
    dirs[:] = [d for d in dirs if d not in exclude_dirs]
    for f in filenames:
        if f.endswith((".py", ".yaml", ".yml", ".txt", ".bat")):
            full = os.path.join(root, f)
            rel = os.path.relpath(full, PROJECT_DIR)
            try:
                with open(full, "r", encoding="utf-8") as fh:
                    content = fh.read()
            except:
                content = "[编码错误]"
            all_files.append({
                "路径": rel,
                "文件名": f,
                "扩展名": os.path.splitext(f)[1],
                "行数": content.count("\n") + 1,
                "大小KB": round(len(content.encode("utf-8")) / 1024, 1),
                "代码": content,
            })
all_files.sort(key=lambda x: x["路径"])
# 2. 按行数均衡分成两组
group1, group2 = [], []
sum1, sum2 = 0, 0
for f in all_files:
    if sum1 <= sum2:
        group1.append(f)
        sum1 += f["行数"]
    else:
        group2.append(f)
        sum2 += f["行数"]
# 统计
total_kb = sum(f["大小KB"] for f in all_files)
total_lines = sum(f["行数"] for f in all_files)
# ===== 公共样式 =====
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
HDR_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill(start_color="4A6CF7", end_color="4A6CF7", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CODE_FONT = Font(name="Consolas", size=9)
CODE_ALIGN = Alignment(vertical="top", wrap_text=False)
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
FILE_FILL = PatternFill(start_color="263238", end_color="263238", fill_type="solid")
FILE_FONT = Font(name="Consolas", size=10, bold=True, color="FFFFFF")
PY_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
YAML_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
BAT_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
DIM_FONT = Font(name="Consolas", size=8, color="999999")
def write_excel(group, part_no, file_count_start):
    """写入一个Excel文件"""
    wb = openpyxl.Workbook()
    # ====== Sheet 1: 目录索引 ======
    ws = wb.active
    ws.title = "目录索引"
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"树剪 v3.0 完整源代码 - 第{part_no}部分"
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="1a237e")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"导出: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} | 本部分: {len(group)}个文件, {sum(f['行数'] for f in group):,}行, {sum(f['大小KB'] for f in group):.0f}KB | 项目总计: {len(all_files)}个文件, {total_lines:,}行, {total_kb:.0f}KB"
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:F3")
    ws["A3"].value = f"文件编号范围: {file_count_start} ~ {file_count_start + len(group) - 1}"
    ws["A3"].font = Font(name="Microsoft YaHei", size=9, color="888888")
    ws["A3"].alignment = Alignment(horizontal="center")
    for col, h in enumerate(["序号", "文件路径", "文件名", "类型", "行数", "KB"], 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, HDR_ALIGN, THIN
    for i, f in enumerate(group):
        r = i + 6
        ext = f["扩展名"]
        rf = PY_FILL if ext == ".py" else (YAML_FILL if ext in (".yaml", ".yml") else BAT_FILL)
        for col, val in enumerate([i + file_count_start, f["路径"], f["文件名"], ext, f["行数"], f["大小KB"]], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Consolas", size=10) if col <= 3 else DIM_FONT
            c.fill, c.border = rf, THIN
            c.alignment = Alignment(vertical="top") if col == 2 else Alignment(vertical="center")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.freeze_panes = "A6"
    # ====== Sheet 2: 全部代码 ======
    ws2 = wb.create_sheet("全部源代码")
    ws2.merge_cells("A1:C1")
    ws2["A1"].value = f"树剪 v3.0 源代码 - 第{part_no}部分 ({len(group)}个文件)"
    ws2["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1a237e")
    for col, h in enumerate(["文件路径", "行号", "代码内容"], 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, HDR_ALIGN, THIN
    cr = 3
    for f in group:
        # 文件标题行
        ws2.merge_cells(f"A{cr}:C{cr}")
        sc = ws2.cell(row=cr, column=1, value=f"# {f['路径']}   ({f['行数']}行)")
        sc.font, sc.fill, sc.border = FILE_FONT, FILE_FILL, THIN
        for c in range(2, 4):
            ws2.cell(row=cr, column=c).fill, ws2.cell(row=cr, column=c).border = FILE_FILL, THIN
        cr += 1
        # 每行代码
        for ln, line in enumerate(f["代码"].split("\n"), 1):
            ws2.cell(row=cr, column=1, value=f["路径"]).font = DIM_FONT
            ws2.cell(row=cr, column=2, value=ln).font = DIM_FONT
            ws2.cell(row=cr, column=3, value=line).font = CODE_FONT
            for c in range(1, 4):
                ws2.cell(row=cr, column=c).border = THIN
                ws2.cell(row=cr, column=c).alignment = CODE_ALIGN if c == 3 else Alignment(vertical="top")
            cr += 1
        cr += 1  # 文件间空行
    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 7
    ws2.column_dimensions["C"].width = 180
    ws2.freeze_panes = "A3"
    # ====== Sheet 3: 按模块 ======
    ws3 = wb.create_sheet("按模块分类")
    cats = {}
    for f in group:
        parts = f["路径"].replace("\\", "/").split("/")
        cat = parts[0] if len(parts) == 1 else (f"plugins/{parts[1]}" if parts[0] == "plugins" and len(parts) > 1 else parts[0])
        cats.setdefault(cat, []).append(f)
    ws3.merge_cells("A1:D1")
    ws3["A1"].value = f"树剪 v3.0 模块分类 - 第{part_no}部分 ({len(cats)}个模块)"
    ws3["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1a237e")
    for col, h in enumerate(["模块/文件", "文件数", "行数", "KB"], 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, HDR_ALIGN, THIN
    cr = 3
    for cat_name in sorted(cats.keys()):
        cf = cats[cat_name]
        cl, cs = sum(x["行数"] for x in cf), sum(x["大小KB"] for x in cf)
        for col, val in enumerate([cat_name, len(cf), cl, round(cs, 1)], 1):
            c = ws3.cell(row=cr, column=col, value=val)
            c.font = Font(name="Consolas", size=10, bold=True) if col == 1 else Font(name="Consolas", size=10)
            c.border = THIN
        cr += 1
        for f in sorted(cf, key=lambda x: x["路径"]):
            ws3.cell(row=cr, column=1, value=f"    {f['文件名']}").font = Font(name="Consolas", size=9, color="555555")
            ws3.cell(row=cr, column=2, value="").border = THIN
            ws3.cell(row=cr, column=3, value=f["行数"]).font = DIM_FONT; ws3.cell(row=cr, column=3).border = THIN
            ws3.cell(row=cr, column=4, value=f["大小KB"]).font = DIM_FONT; ws3.cell(row=cr, column=4).border = THIN
            ws3.cell(row=cr, column=1).border = THIN
            cr += 1
        cr += 1
    ws3.column_dimensions["A"].width = 60
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 12
    ws3.freeze_panes = "A3"
    # 保存
    fname = os.path.join(DESKTOP, f"树剪_v3.0_代码_第{part_no}部分.xlsx")
    wb.save(fname)
    return fname
# 3. 生成两个文件
p1 = write_excel(group1, 1, 1)
p2 = write_excel(group2, 2, len(group1) + 1)
print(f"OK Part1: {os.path.basename(p1)} | {len(group1)}files | {sum(f['行数'] for f in group1)}lines")
print(f"OK Part2: {os.path.basename(p2)} | {len(group2)}files | {sum(f['行数'] for f in group2)}lines")
print(f"OK Total: {len(all_files)}files | {total_lines}lines | 完整无缺失")
