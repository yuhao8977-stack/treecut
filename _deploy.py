# -*- coding: utf-8 -*-
"""v3.0 部署脚本 - 从 Excel 提取全部代码写入项目目录"""
import os, sys
sys.dont_write_bytecode = True

PROJECT_ROOT = r"E:\树剪软件相关文件"
EXCEL1 = r"C:\Users\admin\Desktop\树剪_v3.0_代码_第1部分.xlsx"
EXCEL2 = r"C:\Users\admin\Desktop\树剪_v3.0_代码_第2部分.xlsx"

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q", "-i", "https://pypi.tuna.tsinghua.edu.cn/simple"])
    import openpyxl

def extract_path(line):
    s = line[2:].strip()
    idx = s.rfind("(")
    if idx > 0:
        s = s[:idx].strip()
    return s

def write_file(rel_path, lines):
    full = os.path.join(PROJECT_ROOT, rel_path)
    dn = os.path.dirname(full)
    os.makedirs(dn, exist_ok=True)
    content = "\n".join(lines)
    if not content.endswith("\n"):
        content += "\n"
    with open(full, "w", encoding="utf-8", newline="\n") as f:
        f.write(content)

total = 0
all_paths = []
lines_total = 0

for xlsx_path, label in [(EXCEL1, "Part1"), (EXCEL2, "Part2")]:
    if not os.path.exists(xlsx_path):
        print(f"[{label}] MISSING: {xlsx_path}")
        continue

    print(f"[{label}] Reading {os.path.basename(xlsx_path)}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "全部源代码" not in wb.sheetnames:
        print(f"  ERROR: Sheet '全部源代码' not found. Available: {wb.sheetnames}")
        wb.close()
        continue

    ws = wb["全部源代码"]
    cur_path = None
    cur_lines = []

    for row in ws.iter_rows(min_row=3, max_col=3, values_only=True):
        pv, lv, cv = row
        if pv and isinstance(pv, str) and pv.startswith("# ") and "E:" not in pv:
            if cur_path and cur_lines:
                write_file(cur_path, cur_lines)
                all_paths.append(cur_path)
                lines_total += len(cur_lines)
                total += 1
            cur_path = extract_path(pv)
            cur_lines = []
            continue
        if cur_path and cv is not None:
            cur_lines.append(str(cv))

    if cur_path and cur_lines:
        write_file(cur_path, cur_lines)
        all_paths.append(cur_path)
        lines_total += len(cur_lines)
        total += 1

    wb.close()
    print(f"  [{label}] Done")

print(f"\n{'='*60}")
print(f"DEPLOYED: {total} files, {lines_total} total lines")
print(f"PROJECT_ROOT: {PROJECT_ROOT}")
print(f"{'='*60}")

# Verify critical files
critical = [
    "main.py", "requirements.txt",
    "core/__init__.py", "core/config_loader.py", "core/database.py",
    "core/model_pool.py", "core/plugin_manager.py", "core/process_manager.py",
    "core/workflow_engine.py",
    "plugins/base_plugin.py",
    "plugins/recognize/__init__.py", "plugins/recognize/video_split_plugin.py",
    "plugins/recognize/audio_transcribe_plugin.py", "plugins/recognize/subtitle_ocr_plugin.py",
    "plugins/recognize/scene_tag_plugin.py", "plugins/recognize/bgm_classify_plugin.py",
    "plugins/quality/__init__.py", "plugins/quality/black_frame_plugin.py",
    "plugins/quality/param_check_plugin.py", "plugins/quality/blur_stutter_plugin.py",
    "plugins/quality/aspect_check_plugin.py", "plugins/quality/subtitle_sync_plugin.py",
    "plugins/quality/brand_detect_plugin.py",
    "plugins/correct/__init__.py", "plugins/correct/trim_black_plugin.py",
    "plugins/correct/subtitle_align_plugin.py", "plugins/correct/standard_transcode_plugin.py",
    "utils/__init__.py", "utils/logger.py", "utils/vram_manager.py",
    "utils/system_optimizer.py", "utils/cache_manager.py",
    "utils/ffmpeg_utils.py", "utils/vector_store.py",
    "tasks/__init__.py", "tasks/task_queue.py",
    "dashboard/app.py",
    "config/system_config.yaml", "config/optimize_config.yaml",
    "modules/__init__.py", "modules/video_analyzer.py", "modules/audio_analyzer.py",
    "modules/subtitle_analyzer.py", "modules/advanced_quality.py",
    "modules/scene_understanding.py", "modules/audio_classifier.py",
    "modules/auto_corrector.py",
    "requirements/phase2_optimize.txt",
]

print("\n=== Critical files check ===")
ok = 0
miss = []
for f in critical:
    p = os.path.join(PROJECT_ROOT, f)
    if os.path.exists(p):
        ok += 1
    else:
        miss.append(f)
        print(f"  MISSING: {f}")

print(f"  {ok}/{len(critical)} present")
if miss:
    print(f"  MISSING ({len(miss)}): {', '.join(miss)}")
else:
    print("  ALL CRITICAL FILES PRESENT")

# List all deployed
print(f"\n=== All {total} files ===")
for p in sorted(all_paths):
    print(f"  {p}")
